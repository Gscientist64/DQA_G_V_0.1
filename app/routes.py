# app/routes.py
# FULL VERSION WITH ALL ORIGINAL CODE + FIXES + NEW THEMATIC ANALYSIS ENDPOINT

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, Response, session
import os
import uuid
import logging
import json
from werkzeug.utils import secure_filename
from datetime import datetime
import traceback
from .analysis import (
    process_facility_bundle, 
    process_multiple_facilities, 
    export_results_to_excel,
    analyze_dvv_thematic_areas,
    compare_multiple_dvv_thematic,
    export_thematic_analysis
)
from .storage import (
    load_results, save_results, get_facility_by_name, 
    delete_facility as storage_delete_facility, get_summary_statistics,
    clear_all_results, export_results_to_json
)

STANDALONE_MODE = False

bp = Blueprint("routes", __name__)
logger = logging.getLogger(__name__)

UPLOAD_DIR = os.path.join("data", "uploaded_files")
ALLOWED_EXTENSIONS = {"xls", "xlsx"}
MAX_CONTENT_LENGTH = 50 * 1024 * 1024  # 50MB max file size

# Load existing analyses from disk on startup
def safe_load_results():
    try:
        from .storage import load_results
        results = load_results()
        
        if isinstance(results, list):
            return results
        else:
            logger.error(f"load_results() didn't return a list, got: {type(results)}")
            return []
    except Exception as e:
        logger.error(f"Error loading results: {e}")
        return []

facilities = safe_load_results()

QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

def allowed_file(filename: str) -> bool:
    """Check if file has allowed extension and reasonable size."""
    if not filename or '.' not in filename:
        return False
    
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS


@bp.route("/")
def welcome():
    """Welcome page."""
    logger.info("Accessed welcome page")
    return render_template("welcome.html")


@bp.route("/overview")
def overview():
    """System overview page."""
    logger.info("Accessed overview page")
    return render_template("overview.html")


@bp.route("/manual")
def manual():
    """User manual page."""
    logger.info("Accessed manual page")
    return render_template("manual.html")


@bp.route("/analysis")
def analysis():
    """Analysis upload page - disabled in standalone mode."""
    if STANDALONE_MODE:
        flash("Analysis features are disabled in standalone mode.", "warning")
        return redirect(url_for("routes.performance"))
    
    logger.info("Accessed analysis page")
    stats = get_summary_statistics()
    
    return render_template(
        "analysis.html", 
        facilities=facilities, 
        quarters=QUARTERS,
        stats=stats
    )

@bp.route("/dvv-thematic")
def dvv_thematic():
    """DVV thematic area analysis page."""
    logger.info("Accessed DVV thematic analysis page")
    
    stats = get_summary_statistics()
    
    return render_template(
        "dvv_thematic.html", 
        facilities=facilities, 
        quarters=QUARTERS,
        stats=stats
    )


@bp.route("/batch-upload")
def batch_upload():
    """Batch upload page for multiple facilities."""
    logger.info("Accessed batch upload page")
    return render_template("batch_upload.html", quarters=QUARTERS)


@bp.route("/performance")
def performance():
    """Performance dashboard page."""
    selected_quarter = request.args.get("quarter", "").strip()

    visible_facilities = facilities
    
    if selected_quarter and selected_quarter in QUARTERS:
        quarter_filtered = [f for f in facilities if f.get("quarter") == selected_quarter]
        logger.debug(f"Filtering for quarter {selected_quarter}: {len(quarter_filtered)} facilities")
    else:
        quarter_filtered = facilities

    visible_facilities = sorted(
        quarter_filtered,
        key=lambda f: f.get("Score", 0.0),
        reverse=False,
    )

    stats = get_summary_statistics()
    
    ratings_dist = stats.get("ratings_distribution", {})
    
    logger.info(f"Performance page: {len(visible_facilities)} facilities displayed")
    
    return render_template(
        "performance.html",
        facilities=visible_facilities,
        all_facilities=quarter_filtered,
        quarters=QUARTERS,
        selected_quarter=selected_quarter or "",
        stats=stats,
        ratings_dist=ratings_dist
    )

@bp.route("/results")
def results():
    """Results and winners page."""
    logger.info("Accessed results page")
    
    all_facilities = safe_load_results()
    
    if not all_facilities or len(all_facilities) == 0:
        logger.info("No facilities found for results page")
        return render_template(
            "results.html", 
            all_facilities=[], 
            overall_winner=None, 
            category_winners={}, 
            quarters=QUARTERS,
            stats=get_summary_statistics()
        )

    try:
        valid_facilities = []
        for i, facility in enumerate(all_facilities):
            if isinstance(facility, dict) and "Score" in facility:
                valid_facilities.append(facility)
            else:
                logger.warning(f"Facility at index {i} missing Score or not a dict")
        
        if not valid_facilities:
            logger.info("No facilities with valid scores")
            return render_template(
                "results.html", 
                all_facilities=[], 
                overall_winner=None, 
                category_winners={}, 
                quarters=QUARTERS,
                stats=get_summary_statistics()
            )
        
        logger.info(f"Processing {len(valid_facilities)} valid facilities")
        
        overall_winner = max(valid_facilities, key=lambda x: x.get("Score", 0))
        logger.info(f"Overall winner: {overall_winner.get('Facility')} with score {overall_winner.get('Score')}")

        def safe_max(key):
            items = [f for f in valid_facilities if f.get(key) is not None]
            if not items:
                return None
            return max(items, key=lambda x: x.get(key, 0))

        category_winners = {
            "DVV": safe_max("DVV"),
            "Client_Level": safe_max("Client_Level"),
            "VL_Unsuppressed": safe_max("VL_Unsuppressed"),
            "Availability": safe_max("Availability"),
            "Integrity": safe_max("Integrity"),
            "Consistency": safe_max("Consistency"),
        }
        
        category_winners = {k: v for k, v in category_winners.items() if v is not None}
        
        stats = get_summary_statistics()
        
        if stats and "last_updated" in stats and stats["last_updated"]:
            try:
                from datetime import datetime
                dt = datetime.fromtimestamp(stats["last_updated"])
                stats["last_updated_formatted"] = dt.strftime("%Y-%m-%d")
            except Exception as e:
                logger.warning(f"Could not format timestamp: {e}")
                stats["last_updated_formatted"] = str(stats["last_updated"])
        else:
            stats["last_updated_formatted"] = "N/A"

        logger.info(f"Rendering results page with {len(valid_facilities)} facilities")
        
        return render_template(
            "results.html",
            all_facilities=valid_facilities,
            overall_winner=overall_winner,
            category_winners=category_winners,
            quarters=QUARTERS,
            stats=stats
        )
        
    except Exception as e:
        logger.error(f"Error in results page: {e}", exc_info=True)
        return render_template(
            "results.html", 
            all_facilities=[], 
            overall_winner=None, 
            category_winners={}, 
            quarters=QUARTERS,
            stats=get_summary_statistics()
        )

@bp.route("/upload", methods=["POST"])
def upload_file():
    """Handle single facility upload."""
    logger.info("Processing single facility upload")
    
    if request.content_length and request.content_length > MAX_CONTENT_LENGTH:
        flash(f"Total upload size exceeds {MAX_CONTENT_LENGTH//(1024*1024)}MB limit.")
        logger.warning("Upload size exceeded limit")
        return redirect(url_for("routes.analysis"))

    dvv_file = request.files.get("dvv_file")
    client_file = request.files.get("client_file")
    vl_file = request.files.get("vl_file")

    if not dvv_file or not client_file or not vl_file:
        flash("Please upload all 3 required files: DVV Template, Client Level, and VL Unsuppressed.")
        logger.warning("Missing required files")
        return redirect(url_for("routes.analysis"))

    if dvv_file.filename == "" or client_file.filename == "" or vl_file.filename == "":
        flash("One or more selected files are empty.")
        logger.warning("Empty file submitted")
        return redirect(url_for("routes.analysis"))

    file_checks = [
        (dvv_file, "DVV Template"),
        (client_file, "Client Level"),
        (vl_file, "VL Unsuppressed")
    ]
    
    for file, file_type in file_checks:
        if not allowed_file(file.filename):
            flash(f"Invalid file type for {file_type}. Allowed types: xls, xlsx")
            logger.warning(f"Invalid file type: {file.filename}")
            return redirect(url_for("routes.analysis"))

    quarter = request.form.get("quarter") or "Q4"
    if quarter not in QUARTERS:
        quarter = "Q4"
        logger.debug(f"Invalid quarter selected, defaulting to {quarter}")

    upload_dir = os.path.join("data", "uploaded_files")
    os.makedirs(upload_dir, exist_ok=True)

    bundle_id = str(uuid.uuid4())[:8]

    dvv_name = secure_filename(f"{bundle_id}__{dvv_file.filename}")
    client_name = secure_filename(f"{bundle_id}__{client_file.filename}")
    vl_name = secure_filename(f"{bundle_id}__{vl_file.filename}")

    dvv_path = os.path.join(upload_dir, dvv_name)
    client_path = os.path.join(upload_dir, client_name)
    vl_path = os.path.join(upload_dir, vl_name)

    try:
        logger.info(f"Saving uploaded files: {dvv_name}, {client_name}, {vl_name}")
        
        dvv_file.save(dvv_path)
        client_file.save(client_path)
        vl_file.save(vl_path)
        
        logger.info("Files saved successfully, starting analysis")

        res = process_facility_bundle(dvv_name, client_name, vl_name)
        res["id"] = str(uuid.uuid4())
        res["quarter"] = quarter
        res["upload_timestamp"] = datetime.now().isoformat()
        res["bundle_id"] = bundle_id

        global facilities
        facilities.append(res)
        save_results(facilities)

        logger.info(f"Successfully processed facility: {res.get('Facility')}, Score: {res.get('Score')}")
        
        if "warnings" in res and res["warnings"]:
            warning_msg = f"Processed with warnings: {', '.join(res['warnings'][:3])}"
            flash(warning_msg, "warning")
        else:
            flash(f"Successfully processed {res.get('Facility')} with score {res.get('Score')}%.")

    except Exception as e:
        logger.error(f"Error processing bundle: {str(e)}", exc_info=True)
        flash(f"Error processing bundle: {str(e)}")
        
        for p in [dvv_path, client_path, vl_path]:
            if os.path.exists(p):
                try:
                    os.remove(p)
                    logger.debug(f"Cleaned up file: {p}")
                except Exception as cleanup_error:
                    logger.error(f"Failed to cleanup file {p}: {cleanup_error}")

    return redirect(url_for("routes.performance", quarter=quarter))


@bp.route("/upload-dvv-thematic", methods=["POST"])
def upload_dvv_thematic():
    """Handle DVV file upload for thematic area analysis."""
    logger.info("Processing DVV thematic analysis upload")
    
    if request.content_length and request.content_length > MAX_CONTENT_LENGTH:
        flash(f"File size exceeds {MAX_CONTENT_LENGTH//(1024*1024)}MB limit.")
        logger.warning("DVV thematic upload size exceeded limit")
        return redirect(url_for("routes.dvv_thematic"))

    dvv_file = request.files.get("dvv_file")

    if not dvv_file:
        flash("Please upload a DVV Template file.")
        logger.warning("No DVV file submitted")
        return redirect(url_for("routes.dvv_thematic"))

    if dvv_file.filename == "":
        flash("Selected file is empty.")
        logger.warning("Empty DVV file submitted")
        return redirect(url_for("routes.dvv_thematic"))

    if not allowed_file(dvv_file.filename):
        flash("Invalid file type. Allowed types: xls, xlsx")
        logger.warning(f"Invalid DVV file type: {dvv_file.filename}")
        return redirect(url_for("routes.dvv_thematic"))

    quarter = request.form.get("quarter") or "Q4"
    if quarter not in QUARTERS:
        quarter = "Q4"
        logger.debug(f"Invalid quarter selected, defaulting to {quarter}")

    upload_dir = os.path.join("data", "uploaded_files")
    os.makedirs(upload_dir, exist_ok=True)

    file_id = str(uuid.uuid4())[:8]
    dvv_name = secure_filename(f"{file_id}__{dvv_file.filename}")
    dvv_path = os.path.join(upload_dir, dvv_name)

    try:
        logger.info(f"Saving DVV file for thematic analysis: {dvv_name}")
        dvv_file.save(dvv_path)
        
        logger.info("DVV file saved, starting thematic analysis")
        
        thematic_analysis = analyze_dvv_thematic_areas(dvv_name)
        
        thematic_analysis["analysis_id"] = str(uuid.uuid4())
        thematic_analysis["quarter"] = quarter
        thematic_analysis["upload_timestamp"] = datetime.now().isoformat()
        thematic_analysis["dvv_filename"] = dvv_name
        
        session['current_thematic_analysis'] = thematic_analysis
        
        logger.info(f"Successfully analyzed thematic areas for {thematic_analysis['facility_info']['facility']}")
        flash(f"Thematic analysis complete for {thematic_analysis['facility_info']['facility']}")
        
        return redirect(url_for("routes.dvv_thematic_results", analysis_id=thematic_analysis["analysis_id"]))
        
    except Exception as e:
        logger.error(f"Error processing DVV thematic analysis: {str(e)}", exc_info=True)
        flash(f"Error processing DVV thematic analysis: {str(e)}")
        
        if os.path.exists(dvv_path):
            try:
                os.remove(dvv_path)
                logger.debug(f"Cleaned up file: {dvv_path}")
            except Exception as cleanup_error:
                logger.error(f"Failed to cleanup file {dvv_path}: {cleanup_error}")
        
        return redirect(url_for("routes.dvv_thematic"))


@bp.route("/dvv-thematic-results")
@bp.route("/dvv-thematic-results/<analysis_id>")
def dvv_thematic_results(analysis_id=None):
    """Display thematic analysis results."""
    logger.info(f"Displaying thematic analysis results for ID: {analysis_id}")
    
    thematic_analysis = session.get('current_thematic_analysis')
    
    if not thematic_analysis:
        flash("Thematic analysis not found or session expired.")
        logger.warning("Thematic analysis not found in session")
        return redirect(url_for("routes.dvv_thematic"))
    
    if analysis_id and thematic_analysis.get("analysis_id") != analysis_id:
        flash("Thematic analysis ID mismatch.")
        logger.warning(f"Thematic analysis ID {analysis_id} doesn't match session")
        return redirect(url_for("routes.dvv_thematic"))
    
    return render_template("dvv_thematic_results.html", analysis=thematic_analysis)


@bp.route("/compare-dvv-thematic", methods=["POST"])
def compare_dvv_thematic():
    """Compare thematic areas across multiple DVV files."""
    logger.info("Processing DVV thematic comparison")
    
    if request.content_length and request.content_length > MAX_CONTENT_LENGTH * 5:
        flash("Upload size too large.")
        logger.warning("DVV comparison upload size exceeded limit")
        return redirect(url_for("routes.dvv_thematic"))

    files = request.files.getlist("dvv_files")
    
    if not files:
        flash("No files uploaded.")
        return redirect(url_for("routes.dvv_thematic"))
    
    upload_dir = os.path.join("data", "uploaded_files")
    os.makedirs(upload_dir, exist_ok=True)
    
    facility_files = []
    saved_files = []
    
    try:
        for i, file in enumerate(files):
            if not allowed_file(file.filename):
                flash(f"Invalid file type: {file.filename}")
                raise ValueError(f"Invalid file type: {file.filename}")
            
            file_id = str(uuid.uuid4())[:8]
            file_name = secure_filename(f"{file_id}__{file.filename}")
            file_path = os.path.join(upload_dir, file_name)
            
            file.save(file_path)
            saved_files.append(file_path)
            
            facility_name = os.path.splitext(file.filename)[0]
            facility_files.append({
                "name": facility_name,
                "dvv_file": file_name
            })
        
        logger.info(f"Processing {len(facility_files)} DVV files for comparison")
        
        comparison_results = compare_multiple_dvv_thematic(facility_files)
        
        session['current_comparison_results'] = comparison_results
        
        flash(f"Comparison complete. Processed {comparison_results['facilities_count']} facilities.")
        
        return redirect(url_for("routes.dvv_comparison_results"))
        
    except Exception as e:
        logger.error(f"Error comparing DVV files: {str(e)}", exc_info=True)
        flash(f"Error comparing DVV files: {str(e)}")
        
        for file_path in saved_files:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except:
                    pass
        
        return redirect(url_for("routes.dvv_thematic"))


@bp.route("/dvv-comparison-results")
def dvv_comparison_results():
    """Display DVV comparison results."""
    logger.info("Displaying DVV comparison results")
    
    comparison_results = session.get('current_comparison_results')
    
    if not comparison_results:
        flash("Comparison results not found.")
        logger.warning("Comparison results not found in session")
        return redirect(url_for("routes.dvv_thematic"))
    
    return render_template("dvv_comparison_results.html", results=comparison_results)


@bp.route("/export-thematic-excel")
def export_thematic_excel():
    """Export thematic analysis to Excel."""
    logger.info("Exporting thematic analysis to Excel")
    
    try:
        thematic_analysis = session.get('current_thematic_analysis')
        
        if not thematic_analysis:
            flash("No thematic analysis to export.")
            return redirect(url_for("routes.dvv_thematic"))
        
        output_path = export_thematic_analysis(thematic_analysis)
        
        facility_name = thematic_analysis['facility_info']['facility'].replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f"{facility_name}_thematic_analysis_{timestamp}.xlsx"
        
        logger.info(f"Thematic Excel export created: {output_path}")
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Thematic Excel export failed: {str(e)}", exc_info=True)
        flash(f"Export failed: {str(e)}")
        return redirect(url_for("routes.dvv_thematic"))


@bp.route("/batch-upload", methods=["POST"])
def upload_batch():
    """Handle batch upload of multiple facilities."""
    logger.info("Processing batch upload")
    
    if request.content_length and request.content_length > MAX_CONTENT_LENGTH * 5:
        flash("Batch upload size too large.")
        logger.warning("Batch upload size exceeded limit")
        return redirect(url_for("routes.batch_upload"))

    files = request.files.getlist("facility_files")
    quarter = request.form.get("quarter") or "Q4"
    
    if not files:
        flash("No files uploaded.")
        return redirect(url_for("routes.batch_upload"))
    
    if len(files) % 3 != 0:
        flash("Please upload files in sets of 3 (DVV, Client, VL) for each facility.")
        return redirect(url_for("routes.batch_upload"))
    
    upload_dir = os.path.join("data", "uploaded_files")
    os.makedirs(upload_dir, exist_ok=True)
    
    facility_bundles = []
    saved_files = []
    
    try:
        for i in range(0, len(files), 3):
            dvv_file = files[i]
            client_file = files[i + 1]
            vl_file = files[i + 2]
            
            for file, file_type in [(dvv_file, "DVV"), (client_file, "Client"), (vl_file, "VL")]:
                if not allowed_file(file.filename):
                    flash(f"Invalid file type for {file_type}: {file.filename}")
                    raise ValueError(f"Invalid file type: {file.filename}")
            
            bundle_id = str(uuid.uuid4())[:8]
            
            dvv_name = secure_filename(f"{bundle_id}__{dvv_file.filename}")
            client_name = secure_filename(f"{bundle_id}__{client_file.filename}")
            vl_name = secure_filename(f"{bundle_id}__{vl_file.filename}")
            
            dvv_path = os.path.join(upload_dir, dvv_name)
            client_path = os.path.join(upload_dir, client_name)
            vl_path = os.path.join(upload_dir, vl_name)
            
            dvv_file.save(dvv_path)
            client_file.save(client_path)
            vl_file.save(vl_path)
            
            saved_files.extend([dvv_path, client_path, vl_path])
            
            facility_bundles.append({
                "dvv": dvv_name,
                "client_level": client_name,
                "vl_unsuppressed": vl_name,
                "name": f"Facility_{i//3 + 1}"
            })
        
        logger.info(f"Processing {len(facility_bundles)} facility bundles")
        
        result = process_multiple_facilities(facility_bundles)
        
        global facilities
        
        for res in result["successful"]:
            res["id"] = str(uuid.uuid4())
            res["quarter"] = quarter
            res["upload_timestamp"] = datetime.now().isoformat()
            facilities.append(res)
        
        save_results(facilities)
        
        success_count = len(result["successful"])
        fail_count = len(result["failed"])
        
        flash(f"Batch processing complete. Success: {success_count}, Failed: {fail_count}")
        
        if fail_count > 0:
            error_details = "\n".join([f"{e['facility']}: {e['error']}" for e in result["failed"][:5]])
            flash(f"Failures:\n{error_details}", "warning")
    
    except Exception as e:
        logger.error(f"Batch upload error: {str(e)}", exc_info=True)
        flash(f"Batch upload failed: {str(e)}")
        
        for file_path in saved_files:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except:
                    pass
    
    return redirect(url_for("routes.performance", quarter=quarter))


@bp.route("/delete/<facility_id>", methods=["POST"])
def delete_facility(facility_id):
    """Delete a specific facility analysis."""
    logger.info(f"Deleting facility with ID: {facility_id}")
    
    global facilities
    before = len(facilities)
    
    facility_to_delete = None
    for fac in facilities:
        if fac.get("id") == facility_id:
            facility_to_delete = fac
            break
    
    if facility_to_delete:
        success = storage_delete_facility(facility_to_delete.get("Facility"))
        
        if success:
            facilities = safe_load_results()
            flash("Facility analysis deleted successfully.")
            logger.info(f"Deleted facility ID: {facility_id}")
        else:
            flash("Failed to delete facility from storage.")
            logger.warning(f"Failed to delete facility ID {facility_id} from storage")
    else:
        flash("Facility not found.")
        logger.warning(f"Facility ID {facility_id} not found for deletion")

    q = request.args.get("quarter", "")
    return redirect(url_for("routes.performance", quarter=q))


@bp.route("/facility/<facility_id>")
def facility_detail(facility_id):
    """Display detailed view of a specific facility."""
    logger.info(f"Accessing facility detail for ID: {facility_id}")
    
    facility = next((f for f in facilities if f.get("id") == facility_id), None)

    if not facility:
        try:
            all_facilities = load_results()
            facility = next((f for f in all_facilities if f.get("id") == facility_id), None)
        except Exception as e:
            logger.error(f"Error loading facility details: {e}")
            facility = None

    if not facility:
        flash("Facility not found.")
        logger.warning(f"Facility ID {facility_id} not found")
        return redirect(url_for("routes.performance"))

    logger.info(f"Displaying details for facility: {facility.get('Facility')}")
    return render_template("facility_detail.html", facility=facility)


# ============================================
# NEW API ENDPOINT: Thematic Analysis Details
# ============================================

@bp.route("/api/thematic-analysis/<category>")
def get_thematic_analysis(category):
    """Get thematic analysis data for a specific category (ART, HTS, etc.)"""
    try:
        logger.info(f"Fetching thematic analysis data for category: {category}")
        
        # Load fresh data
        all_facilities = load_results()
        
        if not all_facilities:
            return jsonify({
                "success": False,
                "error": "No facility data available"
            })
        
        # Filter and process facilities for the selected category
        category_data = []
        total_indicator_scores = {}
        
        for facility in all_facilities:
            try:
                facility_name = facility.get("Facility", "Unknown")
                
                # Check if facility has thematic_kpis
                if "thematic_kpis" in facility and facility["thematic_kpis"]:
                    thematic_kpis = facility["thematic_kpis"]
                    
                    if isinstance(thematic_kpis, str):
                        try:
                            thematic_kpis = json.loads(thematic_kpis)
                        except json.JSONDecodeError:
                            continue
                    
                    if category in thematic_kpis:
                        category_info = thematic_kpis[category]
                        avg_score = category_info.get("average", 0)
                        indicators = category_info.get("indicators", [])
                        indicator_scores = category_info.get("scores", [])
                        
                        # Create facility data
                        facility_data = {
                            "facility": facility_name,
                            "quarter": facility.get("quarter", "Unknown"),
                            "score": avg_score,
                            "facility_score": facility.get("Score", 0),
                            "indicators": indicators,
                            "indicator_scores": indicator_scores,
                            "facility_id": facility.get("id", "unknown"),
                            "indicator_details": []
                        }
                        
                        # Add indicator details
                        for i, indicator in enumerate(indicators):
                            if i < len(indicator_scores):
                                facility_data["indicator_details"].append({
                                    "name": indicator,
                                    "score": indicator_scores[i],
                                    "category": category
                                })
                                
                                # Accumulate for overall averages
                                if indicator not in total_indicator_scores:
                                    total_indicator_scores[indicator] = []
                                total_indicator_scores[indicator].append(indicator_scores[i])
                        
                        category_data.append(facility_data)
                
                # Check thematic_areas as fallback
                elif "thematic_areas" in facility and facility["thematic_areas"]:
                    thematic_areas = facility["thematic_areas"]
                    
                    if isinstance(thematic_areas, str):
                        try:
                            thematic_areas = json.loads(thematic_areas)
                        except json.JSONDecodeError:
                            continue
                    
                    # Filter areas by category
                    category_areas = []
                    for area in thematic_areas:
                        if area.get("group") == category:
                            category_areas.append(area)
                    
                    if category_areas:
                        # Calculate average for this category in this facility
                        avg_score = sum(area.get("overall_pct", 0) for area in category_areas) / len(category_areas)
                        
                        facility_data = {
                            "facility": facility_name,
                            "quarter": facility.get("quarter", "Unknown"),
                            "score": avg_score,
                            "facility_score": facility.get("Score", 0),
                            "indicators": [area.get("name", "Unknown") for area in category_areas],
                            "indicator_scores": [area.get("overall_pct", 0) for area in category_areas],
                            "facility_id": facility.get("id", "unknown"),
                            "indicator_details": []
                        }
                        
                        # Add indicator details
                        for area in category_areas:
                            facility_data["indicator_details"].append({
                                "name": area.get("name", "Unknown"),
                                "score": area.get("overall_pct", 0),
                                "category": category
                            })
                            
                            # Accumulate for overall averages
                            indicator_name = area.get("name", "Unknown")
                            if indicator_name not in total_indicator_scores:
                                total_indicator_scores[indicator_name] = []
                            total_indicator_scores[indicator_name].append(area.get("overall_pct", 0))
                        
                        category_data.append(facility_data)
                        
            except Exception as e:
                logger.warning(f"Error processing facility {facility.get('Facility', 'Unknown')} for category {category}: {e}")
                continue
        
        if not category_data:
            return jsonify({
                "success": True,
                "category": category,
                "overall_average": 0,
                "facilities_count": 0,
                "indicator_count": 0,
                "facilities": [],
                "indicator_averages": {}
            })
        
        # Calculate overall average for the category
        overall_avg = sum(item["score"] for item in category_data) / len(category_data)
        
        # Calculate indicator averages across all facilities
        indicator_averages = {}
        for indicator, scores in total_indicator_scores.items():
            if scores:
                indicator_averages[indicator] = sum(scores) / len(scores)
        
        # Sort facilities by score (descending)
        category_data.sort(key=lambda x: x["score"], reverse=True)
        
        logger.info(f"Thematic analysis for {category}: {len(category_data)} facilities, average: {overall_avg:.1f}%")
        
        return jsonify({
            "success": True,
            "category": category,
            "overall_average": round(overall_avg, 2),
            "facilities_count": len(category_data),
            "indicator_count": len(indicator_averages),
            "facilities": category_data[:20],  # Limit to top 20 for performance
            "indicator_averages": indicator_averages
        })
        
    except Exception as e:
        logger.error(f"Error in thematic analysis API for category {category}: {e}", exc_info=True)
        return jsonify({
            "success": False,
            "error": str(e),
            "category": category
        })


@bp.route("/api/thematic-data")
def thematic_data():
    """API endpoint for thematic area data (for charts)."""
    logger.info("API request for thematic data")
    
    try:
        facilities_data = load_results()
    except Exception as e:
        logger.error(f"Error loading facilities for thematic data: {e}")
        return jsonify({})

    thematic_aggregate = {}

    for facility in facilities_data:
        for area in facility.get("thematic_areas", []):
            group = area.get("group")
            name = area.get("name")
            score = area.get("overall_pct")

            if group not in thematic_aggregate:
                thematic_aggregate[group] = {}

            if name not in thematic_aggregate[group]:
                thematic_aggregate[group][name] = []

            thematic_aggregate[group][name].append(score)

    thematic_averages = {}
    for group, indicators in thematic_aggregate.items():
        thematic_averages[group] = {}
        for indicator, scores in indicators.items():
            thematic_averages[group][indicator] = sum(scores) / len(scores) if scores else 0

    logger.debug(f"Thematic data prepared: {len(thematic_averages)} groups")
    return jsonify(thematic_averages)


@bp.route("/api/dvv-thematic-data")
def dvv_thematic_data():
    """API endpoint for DVV thematic area data (for charts)."""
    logger.info("API request for DVV thematic data")
    
    thematic_analysis = session.get('current_thematic_analysis')
    
    if not thematic_analysis:
        return jsonify({"error": "No thematic analysis available"})
    
    chart_data = {
        "categories": {},
        "components": {},
        "performance": {}
    }
    
    if 'category_summary' in thematic_analysis:
        for category, data in thematic_analysis['category_summary'].items():
            chart_data["categories"][category] = data['category_percentage']
    
    if 'component_analysis' in thematic_analysis:
        for component, data in thematic_analysis['component_analysis'].items():
            chart_data["components"][component] = data['percentage']
    
    if 'performance_summary' in thematic_analysis:
        for item in thematic_analysis['performance_summary']:
            chart_data["performance"][item['performance_rating']] = item['count']
    
    return jsonify(chart_data)


@bp.route("/api/facility-stats")
def facility_stats():
    """API endpoint for facility statistics."""
    logger.info("API request for facility stats")
    
    stats = get_summary_statistics()
    return jsonify(stats)


@bp.route("/export/excel")
def export_excel():
    """Export all results to Excel."""
    logger.info("Exporting results to Excel")
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join("data", f"facility_results_{timestamp}.xlsx")
        
        facilities_data = load_results()
        export_results_to_excel(facilities_data, output_path)
        
        logger.info(f"Excel export created: {output_path}")
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f"facility_results_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Excel export failed: {str(e)}", exc_info=True)
        flash(f"Export failed: {str(e)}")
        return redirect(url_for("routes.performance"))


@bp.route("/export/json")
def export_json():
    """Export all results to JSON."""
    logger.info("Exporting results to JSON")
    
    try:
        output_path = export_results_to_json()
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=os.path.basename(output_path),
            mimetype="application/json"
        )
    except Exception as e:
        logger.error(f"JSON export failed: {str(e)}", exc_info=True)
        flash(f"Export failed: {str(e)}")
        return redirect(url_for("routes.performance"))


@bp.route("/clear-all", methods=["POST"])
def clear_all():
    """Clear all facility data."""
    logger.warning("Clearing all facility data")
    
    if request.form.get("confirmation") != "DELETE_ALL":
        flash("Invalid confirmation code.")
        return redirect(url_for("routes.performance"))
    
    success = clear_all_results()
    
    if success:
        global facilities
        facilities = []
        flash("All data cleared successfully.")
        logger.info("All data cleared")
    else:
        flash("Failed to clear data.")
        logger.error("Failed to clear all data")
    
    return redirect(url_for("routes.analysis"))


@bp.route("/api/debug/facility/<facility_id>")
def debug_facility(facility_id):
    """Debug endpoint to view raw facility data (JSON)."""
    logger.debug(f"Debug request for facility ID: {facility_id}")
    
    facility = next((f for f in facilities if f.get("id") == facility_id), None)
    
    if not facility:
        try:
            all_facilities = load_results()
            facility = next((f for f in all_facilities if f.get("id") == facility_id), None)
        except Exception as e:
            logger.error(f"Error loading facility for debug: {e}")
    
    if not facility:
        return jsonify({"error": "Facility not found"}), 404
    
    return Response(
        json.dumps(facility, indent=2, default=str),
        mimetype="application/json"
    )
    

@bp.route("/debug-extraction/<facility_id>")
def debug_extraction(facility_id):
    """Debug route to test DVV extraction directly."""
    try:
        facility = next((f for f in facilities if f.get("id") == facility_id), None)
        
        if not facility:
            flash("Facility not found.")
            return redirect(url_for("routes.performance"))
        
        dvv_filename = facility.get("files", {}).get("dvv")
        if not dvv_filename:
            return jsonify({"error": "DVV file not found for this facility"}), 404
        
        dvv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "uploaded_files", dvv_filename)
        
        if not os.path.exists(dvv_path):
            dvv_path = os.path.join("data", "uploaded_files", dvv_filename)
            if not os.path.exists(dvv_path):
                return jsonify({"error": f"DVV file not found on disk: {dvv_path}"}), 404
        
        import openpyxl
        wb = openpyxl.load_workbook(dvv_path, data_only=True)
        
        if "DVV template" not in wb.sheetnames:
            return jsonify({"error": "DVV template sheet not found"}), 404
        
        ws = wb["DVV template"]
        
        debug_data = {
            "facility_name": facility.get("Facility"),
            "dvv_filename": dvv_filename,
            "file_path": dvv_path,
            "file_exists": os.path.exists(dvv_path),
            "extraction_test": []
        }
        
        for r in range(10, 31):
            dataset = ws.cell(row=r, column=3).value
            description = ws.cell(row=r, column=4).value
            score = ws.cell(row=r, column=43).value
            max_score = ws.cell(row=r, column=44).value
            
            if dataset or score:
                debug_data["extraction_test"].append({
                    "row": r,
                    "dataset": str(dataset) if dataset else "None",
                    "description": str(description)[:50] if description else "None",
                    "score": str(score) if score else "None",
                    "max_score": str(max_score) if max_score else "None"
                })
        
        header_row = None
        for r in range(1, 20):
            cell_val = ws.cell(row=r, column=3).value
            if cell_val and isinstance(cell_val, str) and "Dataset" in cell_val:
                header_row = r
                break
        
        debug_data["header_row"] = header_row
        debug_data["max_row"] = ws.max_row
        debug_data["max_column"] = ws.max_column
        
        return render_template(
            "debug_extraction.html",
            debug_data=debug_data,
            facility=facility
        )
        
    except Exception as e:
        logger.error(f"Error in debug_extraction: {str(e)}", exc_info=True)
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    

@bp.route("/debug-dvv-structure/<facility_id>")
def debug_dvv_structure(facility_id):
    """Debug route to inspect DVV template structure."""
    try:
        facility = next((f for f in facilities if f.get("id") == facility_id), None)
        
        if not facility:
            flash("Facility not found.")
            return redirect(url_for("routes.performance"))
        
        dvv_filename = facility.get("files", {}).get("dvv")
        if not dvv_filename:
            return jsonify({"error": "DVV file not found for this facility"}), 404
        
        dvv_path = os.path.join(UPLOAD_DIR, dvv_filename)
        
        if not os.path.exists(dvv_path):
            return jsonify({"error": "DVV file not found on disk"}), 404
        
        import openpyxl
        wb = openpyxl.load_workbook(dvv_path, data_only=True)
        
        if "DVV template" not in wb.sheetnames:
            return jsonify({"error": "DVV template sheet not found"}), 404
        
        ws = wb["DVV template"]
        
        debug_info = {
            "facility_name": facility.get("Facility"),
            "dvv_filename": dvv_filename,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimensions": f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}", # type: ignore
            "structure_analysis": []
        }
        
        for r in range(1, 31):
            row_data = {}
            for col_idx, col_letter in enumerate(["A", "B", "C", "D", "AQ", "AR"], 1):
                cell = ws[f"{col_letter}{r}"]
                if cell.value:
                    row_data[f"col_{col_letter}"] = {
                        "value": str(cell.value)[:100],
                        "type": type(cell.value).__name__
                    }
            
            if row_data:
                debug_info["structure_analysis"].append({
                    "row": r,
                    "cells": row_data
                })
        
        patterns_found = {
            "dataset_header": False,
            "indicator_start": None,
            "total_row": None
        }
        
        for r in range(1, 20):
            col_c = ws.cell(row=r, column=3).value
            if col_c and isinstance(col_c, str) and "Dataset" in col_c:
                patterns_found["dataset_header"] = r
                break
        
        if patterns_found["dataset_header"]:
            start_row = patterns_found["dataset_header"] + 1
            for r in range(start_row, start_row + 10):
                col_c = ws.cell(row=r, column=3).value
                if col_c and isinstance(col_c, str) and col_c.strip():
                    patterns_found["indicator_start"] = r
                    break
        
        debug_info["patterns"] = patterns_found
        
        return render_template(
            "debug_dvv_structure.html",
            debug_info=debug_info,
            facility=facility
        )
        
    except Exception as e:
        logger.error(f"Error in debug_dvv_structure: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@bp.route("/test-extraction/<facility_id>")
def test_extraction(facility_id):
    """Test extraction directly"""
    try:
        facility = next((f for f in facilities if f.get("id") == facility_id), None)
        
        if not facility:
            return jsonify({"error": "Facility not found"}), 404
        
        dvv_filename = facility.get("files", {}).get("dvv")
        if not dvv_filename:
            return jsonify({"error": "No DVV file"}), 404
        
        from app.analysis import _extract_dvv_from_template
        import os
        
        possible_paths = [
            os.path.join("data", "uploaded_files", dvv_filename),
            os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "uploaded_files", dvv_filename)
        ]
        
        dvv_path = None
        for path in possible_paths:
            if os.path.exists(path):
                dvv_path = path
                break
        
        if not dvv_path:
            return jsonify({"error": f"File not found. Tried: {possible_paths}"}), 404
        
        result = _extract_dvv_from_template(dvv_path)
        
        return jsonify({
            "facility": facility.get("Facility"),
            "file": dvv_filename,
            "path": dvv_path,
            "exists": os.path.exists(dvv_path),
            "extraction_result": {
                "Facility": result.get("Facility"),
                "DVV": result.get("DVV"),
                "thematic_areas_count": len(result.get("thematic_areas", [])),
                "thematic_areas": result.get("thematic_areas", [])[:10],
                "has_thematic_kpis": "thematic_kpis" in result
            }
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "traceback": traceback.format_exc()
        }), 500


@bp.route("/health")
def health_check():
    """Health check endpoint."""
    return jsonify({
        "status": "healthy",
        "facility_count": len(facilities),
        "timestamp": datetime.now().isoformat()
    })


@bp.route("/debug/data-flow")
def debug_data_flow():
    """Debug the data flow from storage to template."""
    from .storage import load_results
    
    raw_data = load_results()
    
    from routes import safe_load_results
    safe_data = safe_load_results()
    
    return jsonify({
        "storage_load_results": {
            "type": type(raw_data).__name__,
            "length": len(raw_data) if isinstance(raw_data, list) else "N/A",
            "sample": raw_data[:1] if isinstance(raw_data, list) and len(raw_data) > 0 else raw_data
        },
        "safe_load_results": {
            "type": type(safe_data).__name__,
            "length": len(safe_data) if isinstance(safe_data, list) else "N/A",
            "sample": safe_data[:1] if isinstance(safe_data, list) and len(safe_data) > 0 else safe_data
        },
        "global_facilities": {
            "type": type(facilities).__name__,
            "length": len(facilities) if isinstance(facilities, list) else "N/A"
        }
    })


# Error handlers
@bp.errorhandler(413)
def too_large(e):
    """Handle file too large errors."""
    logger.warning(f"File too large error: {e}")
    flash(f"File too large. Maximum size is {MAX_CONTENT_LENGTH//(1024*1024)}MB.")
    return redirect(url_for("routes.analysis")), 413


@bp.errorhandler(404)
def not_found(e):
    """Handle 404 errors."""
    logger.warning(f"404 error: {e}")
    return render_template("404.html"), 404


@bp.errorhandler(500)
def server_error(e):
    """Handle 500 errors."""
    logger.error(f"500 error: {e}", exc_info=True)
    return render_template("500.html"), 500


# Update global facilities when module loads
def update_facilities_list():
    """Update the global facilities list from storage."""
    global facilities
    facilities = safe_load_results()


# Initialize
update_facilities_list()
logger.info(f"Routes module initialized with {len(facilities)} facilities")