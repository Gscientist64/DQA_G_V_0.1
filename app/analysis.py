# app/analysis.py - UPDATED VERSION WITH FACILITY FILTERING AND COLUMN FIXES FOR BOTH CLIENT AND VL

import os
import openpyxl
import pandas as pd
import numpy as np
import json
import logging
import time
from functools import lru_cache
from contextlib import contextmanager
from datetime import datetime
import warnings
import re
warnings.filterwarnings('ignore')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(BASE_DIR, "data", "uploaded_files")
CONFIG_PATH = os.path.join(BASE_DIR, "config", "analysis_config.json")

# Setup logging
logger = logging.getLogger(__name__)

# Configuration
ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.xlsm'}

# -----------------------------
# Context Managers and Decorators
# -----------------------------

@contextmanager
def timer(name):
    """Context manager to time operations"""
    start = time.time()
    yield
    elapsed = time.time() - start
    logger.info(f"{name} took {elapsed:.2f} seconds")


# -----------------------------
# Configuration Management
# -----------------------------

def load_config():
    """Load column mappings and settings from config file"""
    default_config = {
        "client_level": {
            "start_col": "Sex",
            "end_col": "Biometric enrollment form available in clients folder.1",
            "data_column": "C",
            "fallback_start_cols": ["sex", "sex.1", "gender", "patient sex"],
            "fallback_end_cols": ["biometric enrollment form available in clients folder.1", 
                                 "biometric enrollment form available", "biometric enrollment"]
        },
        "vl_unsuppressed": {
            "start_col": "Sex ",
            "end_col": "Does the folder contain the VL result form used to commence EAC?",
            "data_column": "C",
            "fallback_start_cols": ["sex ", "sex", "gender", "patient sex"],
            "fallback_end_cols": ["does the folder contain the vl result form used to commence eac?", 
                                "vl result form", "eac form", "commence eac", "folder contain"],
            "column_position_fallback": 13  # P to AB is 13 columns (P=16, AB=28, 28-16+1=13)
        },
        "thresholds": {
            "excellent": 90,
            "good": 75,
            "poor": 50
        },
        "performance": {
            "cache_size": 32,
            "max_sheet_rows": 10000
        },
        # Updated DVV thematic area configuration
        "dvv_thematic": {
            "indicator_column": "A",
            "total_score_column": "AQ",
            "availability_column": "J",
            "integrity_column": "N",
            "consistency_column": "S",
            "validity_column": "AO",
            "header_row": 10,
            "indicator_start_row": 11,
            "exclude_phrases": ["total", "score", "max", "dataset", "facility", "completed", 
                               "designation", "signature", "date", "summary", "action",
                               "grand", "overall", "average", "mean"]
        }
    }
    
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r') as f:
                user_config = json.load(f)
                def merge_dicts(default, user):
                    result = default.copy()
                    for key, value in user.items():
                        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                            result[key] = merge_dicts(result[key], value)
                        else:
                            result[key] = value
                    return result
                
                return merge_dicts(default_config, user_config)
        except Exception as e:
            logger.warning(f"Failed to load config file: {e}. Using defaults.")
    
    return default_config

# Load config once
CONFIG = load_config()

# -----------------------------
# Helpers
# -----------------------------

def _validate_file_extension(file_path):
    """Validate that file has an allowed extension"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file format: {ext}. Allowed: {ALLOWED_EXTENSIONS}")


def _parse_percent_to_pct(value):
    """
    Normalize Excel percent-like values to 0-100 float.
    """
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip().replace("%", "").strip()
        if not s:
            return None
        try:
            num = float(s)
        except Exception:
            return None
        return num
    if isinstance(value, (int, float)):
        num = float(value)
        if 0 <= num <= 1:
            return num * 100.0
        return num
    return None


def _find_total_dqa_row(ws):
    """
    Find 'Total DQA Score' row by searching column D.
    """
    for r in range(70, 130):
        cell_val = ws.cell(row=r, column=4).value
        if isinstance(cell_val, str) and cell_val.strip().lower().startswith("total dqa score"):
            return r
    return 97


def _get_facility_name(ws, fallback_name: str) -> str:
    """
    Facility name is stored in V4.
    """
    try:
        name = ws["V4"].value
        if isinstance(name, str) and name.strip():
            return name.strip()
    except Exception:
        pass
    return fallback_name


def _normalize_header(s):
    """
    Normalize Excel column headers.
    """
    if s is None:
        return ""
    
    s = str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.replace('"', "").replace("'", "")
    s = s.replace("\t", " ").replace("\xa0", " ")
    
    import re
    s = re.sub(r'\([^)]*\)', '', s)
    
    s = s.strip()
    s = s.replace('"', '')
    s = re.sub(r'\s+', ' ', s)
    s = s.lower()
    
    return s


def _find_col_index(df, target_name: str):
    """
    Find the index of a column in df by normalized header match.
    """
    target_norm = _normalize_header(target_name)
    normalized_cols = [_normalize_header(c) for c in df.columns]

    if target_norm in normalized_cols:
        return normalized_cols.index(target_norm)

    raise ValueError(
        f"Column not found: '{target_name}'. "
        f"Closest candidates (normalized): {normalized_cols[:20]} ..."
    )


def _find_col_index_with_fallback(df, target_name: str, fallback_names: list = None):  # type: ignore
    """Try multiple column names if exact match fails"""
    try:
        return _find_col_index(df, target_name)
    except ValueError as e:
        if fallback_names:
            for fallback in fallback_names:
                try:
                    return _find_col_index(df, fallback)
                except ValueError:
                    continue
        
        if target_name == CONFIG["vl_unsuppressed"]["end_col"]:
            column_position = CONFIG["vl_unsuppressed"].get("column_position_fallback")
            if column_position is not None and column_position < len(df.columns):
                logger.warning(f"Using column position fallback {column_position} for '{target_name}'")
                return column_position
        
        raise e


@lru_cache(maxsize=CONFIG["performance"]["cache_size"])
def _read_excel_cached(file_path: str):
    """Cache Excel file reads for performance"""
    logger.debug(f"Reading Excel file (cached): {file_path}")
    
    ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if ext == '.xls':
            try:
                import xlrd
                df = pd.read_excel(file_path, engine='xlrd')
                logger.info(f"Successfully read .xls file with xlrd: {file_path}")
                return df
            except ImportError:
                logger.error("xlrd is not installed. Please install it with: pip install xlrd")
                raise ImportError("xlrd package is required to read .xls files. Install with: pip install xlrd")
            except Exception as e:
                logger.error(f"Error reading .xls file with xlrd: {e}")
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                    logger.info(f"Successfully read .xls file with openpyxl fallback: {file_path}")
                    return df
                except Exception as e2:
                    logger.error(f"Failed to read .xls file with openpyxl: {e2}")
                    raise e
        else:
            return pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path}: {e}")
        raise


def _extract_file_metadata(file_path):
    """Extract metadata about the Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        metadata = {
            "filename": os.path.basename(file_path),
            "sheets": wb.sheetnames,
            "created": os.path.getctime(file_path),
            "modified": os.path.getmtime(file_path),
            "size": os.path.getsize(file_path),
            "size_mb": round(os.path.getsize(file_path) / (1024 * 1024), 2)
        }
        
        if len(wb.sheetnames) > 0:
            ws = wb[wb.sheetnames[0]]
            metadata["dimensions"] = ws.calculate_dimension()
            metadata["max_row"] = ws.max_row
            metadata["max_column"] = ws.max_column
        
        return metadata
    except Exception as e:
        logger.error(f"Failed to extract metadata from {file_path}: {e}")
        return {
            "filename": os.path.basename(file_path),
            "error": str(e)
        }


def _clean_facility_name(name):
    """Clean facility name for matching with better differentiation."""
    if not name or not isinstance(name, str):
        return ""
    
    name = name.strip()
    
    # Remove everything in parentheses first
    name = re.sub(r'\s*\([^)]*\)', '', name)
    
    # Remove everything after dash (but keep important qualifiers)
    # Only remove if dash is followed by generic terms
    dash_match = re.search(r'\s*-\s*(.*)', name)
    if dash_match:
        after_dash = dash_match.group(1).lower()
        # Only remove if after dash contains generic terms (not specific qualifiers)
        generic_terms = ['main', 'branch', 'annex', 'extension', 'satellite', 'unit']
        if any(term in after_dash for term in generic_terms):
            name = name[:dash_match.start()].strip()
    
    # Remove everything after period (but only if it's generic)
    period_match = re.search(r'\s*\.\s*(.*)', name)
    if period_match:
        after_period = period_match.group(1).lower()
        if after_period in ['', 'inc', 'ltd', 'plc']:
            name = name[:period_match.start()].strip()
    
    # Keep specific hospital types that differentiate facilities
    # Don't remove these as they're important for differentiation
    facility_types = [
        'catholic', 'general', 'teaching', 'memorial', 'maternity', 
        'community', 'district', 'regional', 'national', 'specialist',
        'referral', 'primary', 'secondary', 'tertiary'
    ]
    
    # We'll keep the type information for now, but normalize it
    for ftype in facility_types:
        name = name.replace(f' {ftype} ', ' ').replace(f' {ftype.title()} ', ' ')
    
    # Normalize common abbreviations
    replacements = {
        'hosp': 'hospital',
        'hosp.': 'hospital',
        'ctr': 'center',
        'ctr.': 'center',
        'ctre': 'center',
        'clnc': 'clinic',
        'clnc.': 'clinic',
        'gen': 'general',
        'gen.': 'general',
        'cath': 'catholic',
        'cath.': 'catholic',
        'mat': 'maternity',
        'mat.': 'maternity'
    }
    
    for abbr, full in replacements.items():
        name = re.sub(rf'\b{re.escape(abbr)}\b', full, name, flags=re.IGNORECASE)
    
    # Remove extra spaces and normalize
    name = re.sub(r'\s+', ' ', name)
    
    return name.strip().lower()

def _get_facility_keywords(name):
    """
    Extract keywords from facility name for better matching.
    Returns a tuple of (location, type, main_name)
    """
    if not name:
        return ("", "", "")
    
    name_lower = name.lower()
    
    # Define facility types to look for
    facility_types = [
        'catholic', 'general', 'teaching', 'memorial', 'maternity',
        'community', 'district', 'regional', 'national', 'specialist',
        'referral', 'primary', 'secondary', 'tertiary', 'university',
        'federal', 'state', 'local'
    ]
    
    # Define hospital/service types
    service_types = [
        'hospital', 'clinic', 'health center', 'health centre',
        'medical center', 'medical centre', 'dispensary', 'pharmacy',
        'laboratory', 'diagnostic center'
    ]
    
    # Extract facility type
    found_type = ""
    for ftype in facility_types:
        if ftype in name_lower:
            found_type = ftype
            break
    
    # Extract service type
    found_service = ""
    for stype in service_types:
        if stype in name_lower:
            found_service = stype
            break
    
    # Try to extract location (usually first word or two)
    words = name_lower.split()
    location = ""
    if words:
        # Common location indicators
        location_candidates = []
        for word in words[:2]:  # First 1-2 words often indicate location
            if word not in facility_types and word not in service_types:
                location_candidates.append(word)
        
        if location_candidates:
            location = " ".join(location_candidates)
    
    # Extract main name (without location and types)
    main_name = name_lower
    if location:
        main_name = main_name.replace(location, "").strip()
    if found_type:
        main_name = main_name.replace(found_type, "").strip()
    if found_service:
        main_name = main_name.replace(found_service, "").strip()
    
    # Clean up the main name
    main_name = re.sub(r'\s+', ' ', main_name).strip()
    
    return (location, found_type or found_service, main_name)

def _match_facility_name(row_facility, target_facility):
    """Check if row facility name matches target facility name with improved logic."""
    if not row_facility or not target_facility:
        return False
    
    row_str = str(row_facility).strip()
    target_str = str(target_facility).strip()
    
    # Clean names
    row_clean = _clean_facility_name(row_str)
    target_clean = _clean_facility_name(target_str)
    
    # Exact match after cleaning
    if row_clean == target_clean:
        logger.debug(f"Exact match: '{row_clean}' == '{target_clean}'")
        return True
    
    # Extract keywords for both names
    row_location, row_type, row_main = _get_facility_keywords(row_str)
    target_location, target_type, target_main = _get_facility_keywords(target_str)
    
    logger.debug(f"Row: '{row_str}' -> loc: '{row_location}', type: '{row_type}', main: '{row_main}'")
    logger.debug(f"Target: '{target_str}' -> loc: '{target_location}', type: '{target_type}', main: '{target_main}'")
    
    # STRICT MATCHING: For similar names, require strong evidence
    # Check if they have different types but similar main names
    if row_main == target_main and row_main:  # Same main name (e.g., "ogoja")
        # Check if types are explicitly different
        row_has_type = any(word in row_str.lower() for word in ['catholic', 'general', 'maternity', 'teaching'])
        target_has_type = any(word in target_str.lower() for word in ['catholic', 'general', 'maternity', 'teaching'])
        
        if row_has_type and target_has_type:
            # Get the specific type words
            row_type_words = [word for word in ['catholic', 'general', 'maternity', 'teaching'] 
                             if word in row_str.lower()]
            target_type_words = [word for word in ['catholic', 'general', 'maternity', 'teaching'] 
                                if word in target_str.lower()]
            
            # If they have different specific type words, they're different facilities
            if row_type_words and target_type_words and row_type_words != target_type_words:
                logger.debug(f"Different facility types: {row_type_words} vs {target_type_words}")
                return False
    
    # Partial match logic (relaxed for other cases)
    # If one name contains the other (and vice versa) after cleaning
    if target_clean in row_clean and len(target_clean) > 3:
        # But check if it's not just a partial match of common words
        common_words = ['hospital', 'clinic', 'health', 'center', 'centre']
        if any(word in target_clean for word in common_words):
            # Need additional verification
            row_words = set(row_clean.split())
            target_words = set(target_clean.split())
            common = row_words.intersection(target_words)
            
            # Count non-common-words
            non_common_words = [w for w in row_words.union(target_words) 
                               if w not in common_words and len(w) > 2]
            
            if len(common) >= 2 and len(non_common_words) > 0:
                logger.debug(f"Strong partial match with {len(common)} common words")
                return True
        else:
            logger.debug(f"Partial match: '{target_clean}' in '{row_clean}'")
            return True
    
    if row_clean in target_clean and len(row_clean) > 3:
        logger.debug(f"Partial match: '{row_clean}' in '{target_clean}'")
        return True
    
    # Word-by-word matching with similarity scoring
    row_words = set(row_clean.split())
    target_words = set(target_clean.split())
    common_words = row_words.intersection(target_words)
    
    # Remove common insignificant words
    insignificant = {'the', 'and', 'of', 'for', 'in', 'at', 'to', 'on', 'by'}
    common_words = common_words - insignificant
    row_words = row_words - insignificant
    target_words = target_words - insignificant
    
    # Calculate Jaccard similarity
    if row_words and target_words:
        similarity = len(common_words) / len(row_words.union(target_words))
        
        # Require higher similarity for short names
        min_similarity = 0.6 if min(len(row_words), len(target_words)) <= 3 else 0.5
        
        if similarity >= min_similarity:
            logger.debug(f"Jaccard similarity: {similarity:.2f} >= {min_similarity}")
            
            # Additional check: if names are very similar but have different key types
            if similarity >= 0.7:
                # Check for contradictory type indicators
                contradictory_pairs = [
                    ('catholic', 'general'),
                    ('general', 'maternity'),
                    ('teaching', 'district'),
                    ('memorial', 'catholic')
                ]
                
                row_lower = row_str.lower()
                target_lower = target_str.lower()
                
                for type1, type2 in contradictory_pairs:
                    if (type1 in row_lower and type2 in target_lower) or \
                       (type2 in row_lower and type1 in target_lower):
                        logger.debug(f"Contradictory types: {type1} vs {type2}")
                        return False
            
            return True
    
    # Levenshtein distance for fuzzy matching (fallback)
    if len(row_clean) > 5 and len(target_clean) > 5:
        from difflib import SequenceMatcher
        similarity = SequenceMatcher(None, row_clean, target_clean).ratio()
        
        if similarity >= 0.8:
            # Double-check with keyword extraction
            if row_location == target_location and row_location:
                # Same location, check if types are compatible
                compatible_types = [
                    ('', 'hospital'),  # Empty type compatible with hospital
                    ('hospital', 'health center'),
                    ('clinic', 'health center'),
                ]
                
                type_pair = (row_type, target_type)
                type_pair_rev = (target_type, row_type)
                
                if type_pair in compatible_types or type_pair_rev in compatible_types:
                    logger.debug(f"Levenshtein match: {similarity:.2f} with compatible types")
                    return True
    
    logger.debug(f"No match: '{row_str}' vs '{target_str}'")
    return False

def _find_facility_column(df, facility_name):
    """
    Find which column contains facility names in the DataFrame.
    Enhanced with multiple strategies.
    """
    # Common facility name patterns (expanded list)
    facility_patterns = [
        'facility', 'site', 'hospital', 'clinic', 'health', 'name',
        'organization', 'institution', 'centre', 'center', 'location',
        'facility name', 'site name', 'health facility', 'service point'
    ]
    
    strategies = []
    
    # Strategy 1: Column name pattern matching
    for col_idx, col_name in enumerate(df.columns):
        col_str = str(col_name).lower()
        
        # Check if column name suggests it contains facility names
        for pattern in facility_patterns:
            if pattern in col_str:
                logger.info(f"Strategy 1: Found facility column by name pattern: '{col_name}' (index {col_idx})")
                strategies.append((col_idx, 1.0, "name_pattern"))
                break
    
    # Strategy 2: Content analysis
    for col_idx, col_name in enumerate(df.columns):
        if col_idx > 10:  # Limit search to first 10 columns for performance
            break
            
        sample_values = df[col_name].dropna().head(10)
        if len(sample_values) < 5:
            continue
        
        # Analyze sample values
        facility_like_count = 0
        avg_length = 0
        
        for val in sample_values:
            val_str = str(val).strip()
            if not val_str:
                continue
            
            avg_length += len(val_str)
            
            # Check if value looks like a facility name
            is_facility_like = False
            
            # Length check
            if len(val_str) > 3 and len(val_str) < 100:
                # Contains common facility words
                facility_words = ['hospital', 'clinic', 'health', 'centre', 'center', 'dispensary']
                if any(word in val_str.lower() for word in facility_words):
                    is_facility_like = True
                # Contains location indicators (often capital letters or specific patterns)
                elif re.search(r'[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*', val_str):
                    # Has capitalized words (like "Ogoja General")
                    is_facility_like = True
                # Not a date, number, or boolean
                elif not re.match(r'^\d+$', val_str):  # Not just numbers
                    if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}', val_str):  # Not date
                        if val_str.lower() not in ['true', 'false', 'yes', 'no', 'ok', 'na', 'n/a']:
                            # Check if it has multiple words (like facility names often do)
                            if len(val_str.split()) >= 2:
                                is_facility_like = True
            
            if is_facility_like:
                facility_like_count += 1
        
        if len(sample_values) > 0:
            avg_length /= len(sample_values)
        
        # Calculate confidence score
        if facility_like_count >= 3:
            confidence = facility_like_count / len(sample_values)
            if avg_length > 8:  # Longer names are more likely to be facilities
                confidence *= 1.2
            
            if confidence >= 0.5:
                logger.info(f"Strategy 2: Column '{col_name}' looks {confidence:.0%} like facility names")
                strategies.append((col_idx, confidence, "content_analysis"))
    
    # Strategy 3: Check column position (often column C in DQA templates)
    if 2 < len(df.columns):  # Column C exists
        strategies.append((2, 0.3, "default_position"))
    
    # Choose the best strategy
    if strategies:
        # Sort by confidence score
        strategies.sort(key=lambda x: x[1], reverse=True)
        best_col_idx, confidence, strategy = strategies[0]
        
        logger.info(f"Selected column index {best_col_idx} '{df.columns[best_col_idx]}' "
                   f"with {confidence:.0%} confidence using {strategy}")
        
        # Verify with a sample
        sample = df.iloc[:5, best_col_idx].dropna().tolist()
        if sample:
            logger.info(f"Sample facility names in column: {sample[:3]}")
        
        return best_col_idx
    
    # Fallback: try to find any column with string values
    for col_idx, col_name in enumerate(df.columns):
        if col_idx > 15:  # Limit search
            break
            
        sample = df[col_name].dropna().head(5).tolist()
        if len(sample) >= 3:
            string_count = sum(1 for val in sample if isinstance(val, str) and len(str(val).strip()) > 3)
            if string_count >= 3:
                logger.warning(f"Fallback: Using column {col_idx} '{col_name}' by string detection")
                return col_idx
    
    # Ultimate fallback
    logger.warning("Could not determine facility column, using default column C (index 2)")
    return 2 if len(df.columns) > 2 else 0

def _filter_by_facility(df, facility_name):
    """
    Filter DataFrame to include only rows for the specified facility.
    Enhanced with better logging and validation.
    """
    if facility_name is None:
        logger.warning("No facility name provided for filtering")
        return df, None
    
    logger.info(f"=== FILTERING FOR FACILITY: {facility_name} ===")
    
    # Find which column contains facility names
    facility_col_idx = _find_facility_column(df, facility_name)
    facility_col_name = df.columns[facility_col_idx]
    
    logger.info(f"Using column '{facility_col_name}' (index {facility_col_idx}) for facility names")
    
    # Get unique facility names in the column for debugging
    unique_facilities = df[facility_col_name].dropna().unique()
    logger.info(f"Found {len(unique_facilities)} unique facility names in column")
    
    if len(unique_facilities) <= 10:
        logger.info(f"All facility names in column: {list(unique_facilities)}")
    else:
        logger.info(f"First 10 facility names: {list(unique_facilities[:10])}")
    
    # Create a copy to avoid modifying the original
    df_filtered = df.copy()
    
    # Apply matching with detailed logging
    matches = []
    non_matches = []
    
    def matches_facility(row_val):
        if pd.isna(row_val):
            return False
        
        row_str = str(row_val)
        is_match = _match_facility_name(row_str, facility_name)
        
        if is_match:
            matches.append(row_str)
        else:
            non_matches.append(row_str)
        
        return is_match
    
    df_filtered['_facility_match'] = df_filtered[facility_col_name].apply(matches_facility)
    
    # Filter rows
    before_count = len(df_filtered)
    df_filtered = df_filtered[df_filtered['_facility_match'] == True]
    after_count = len(df_filtered)
    
    # Remove the temporary column
    df_filtered = df_filtered.drop(columns=['_facility_match'])
    
    logger.info(f"Facility filtering results:")
    logger.info(f"  Total rows before: {before_count}")
    logger.info(f"  Rows after filtering: {after_count}")
    logger.info(f"  Filtered out: {before_count - after_count} rows")
    
    if matches:
        logger.info(f"  Matched facility names: {list(set(matches))}")
    
    if after_count == 0:
        logger.warning(f"⚠️ No rows found for facility '{facility_name}'")
        logger.warning(f"  Column used: '{facility_col_name}'")
        logger.warning(f"  Sample values in column: {df[facility_col_name].dropna().unique()[:5]}")
        
        # Try to find similar facility names
        similar_names = []
        for unique_name in unique_facilities:
            if isinstance(unique_name, str):
                from difflib import SequenceMatcher
                similarity = SequenceMatcher(None, facility_name.lower(), unique_name.lower()).ratio()
                if similarity > 0.7:
                    similar_names.append((unique_name, similarity))
        
        if similar_names:
            similar_names.sort(key=lambda x: x[1], reverse=True)
            logger.warning(f"  Similar facility names found:")
            for name, sim in similar_names[:3]:
                logger.warning(f"    - '{name}' (similarity: {sim:.1%})")
        
        # Return empty dataframe but with same columns
        return df.iloc[0:0], facility_col_idx
    
    logger.info(f"✓ Successfully found {after_count} rows for facility '{facility_name}'")
    
    # Log a sample of the filtered data
    if after_count > 0:
        sample_data = df_filtered.iloc[:3, max(0, facility_col_idx-2):facility_col_idx+3]
        logger.debug(f"Sample of filtered data (columns around facility column):")
        for idx, row in sample_data.iterrows():
            logger.debug(f"  Row {idx}: {row.tolist()}")
    
    return df_filtered, facility_col_idx

# -----------------------------
# DVV Extract with IMPROVED Thematic Area Categorization
# -----------------------------

def _extract_dvv_from_template(dvv_path: str):
    """
    DVV final score extraction - EXTRACTING SPECIFIC DATASETS ONLY
    Based on Q1 requirements with proper averaging for POST_RESP and TX_ML
    Scores are calculated as: (AQ score / AR max score) * 100%
    """
    _validate_file_extension(dvv_path)
    
    logger.info(f"Extracting DVV from: {dvv_path}")
    
    with timer(f"Loading DVV workbook: {os.path.basename(dvv_path)}"):
        wb = openpyxl.load_workbook(dvv_path, data_only=True)

    if "DVV template" not in wb.sheetnames:
        raise ValueError("DVV template sheet not found (expected: 'DVV template')")

    ws = wb["DVV template"]

    fallback = os.path.splitext(os.path.basename(dvv_path))[0]
    facility_name = _get_facility_name(ws, fallback)
    
    logger.info(f"Processing facility: {facility_name}")

    total_row = _find_total_dqa_row(ws)
    logger.debug(f"Found Total DQA Score at row {total_row}")

    # Overall DVV final score
    dvv_val = ws[f"AS{total_row}"].value
    dvv_pct = _parse_percent_to_pct(dvv_val)
    if dvv_pct is None:
        raise ValueError(f"Could not read DVV final score from AS{total_row}")

    dvv_pct = round(dvv_pct, 2)
    logger.info(f"DVV score: {dvv_pct}%")

    # Availability / Integrity / Consistency / Validity
    def pct_cell(col_letter):
        v = ws[f"{col_letter}{total_row}"].value
        p = _parse_percent_to_pct(v)
        return round(p, 2) if p is not None else None

    availability = pct_cell("J")
    integrity = pct_cell("N")
    consistency = pct_cell("S")
    validity = pct_cell("AO")
    
    logger.debug(f"Availability: {availability}%, Integrity: {integrity}%, Consistency: {consistency}%, Validity: {validity}%")

    # ====================================================
    # SPECIFIC DATASET EXTRACTION FOR Q1
    # ====================================================
    
    thematic_areas = []
    logger.info("Extracting specific Q1 datasets...")
    
    # First, let's find where datasets are by scanning for known patterns
    dataset_rows = {}  # Map dataset name to row number
    
    # Scan through the sheet to find our specific datasets
    for r in range(10, ws.max_row + 1):
        # Check column C for dataset names
        cell_val = ws.cell(row=r, column=3).value  # Column C
        if not cell_val:
            continue
            
        cell_str = str(cell_val).strip().upper()
        
        # Check for our specific datasets
        target_datasets = [
            "HTS_TST", "HTS_TST_POS", "HTS_SELF", "POST_RESP",
            "PMTCT_STAT DENOMINATOR", "PMTCT_STAT NUMERATOR", "PMTCT_STAT_POS",
            "PMTCT_ART (N)", "PMTCT_EID (N)", "PMTCT_HEI_NEG", "PMTCT_HEI_POS",
            "PMTCT_HEI_POS ON ART", "TB_STAT(N)", "TB_STAT(D)", "PREP_NEW",
            "PREP_ CT", "TX_NEW", "TX_CURR", "TX_RTT", "TRANSFER_INS",
            "TX_ML", "TX_PVLS (N)", "TX_PVLS (D)"
        ]
        
        for target in target_datasets:
            if target in cell_str:
                dataset_rows[target] = r
                logger.debug(f"Found '{target}' at row {r}")
                break
    
    logger.info(f"Found {len(dataset_rows)} dataset rows")
    
    # Helper function to extract and calculate percentage
    def calculate_percentage(score_val, max_val):
        """Calculate percentage: (score / max) * 100"""
        # Parse score
        score = None
        if isinstance(score_val, (int, float)):
            score = float(score_val)
        elif isinstance(score_val, str):
            try:
                clean_score = re.sub(r'[^\d.]', '', score_val)
                if clean_score:
                    score = float(clean_score)
            except:
                pass
        
        # Parse max score
        max_score = None
        if isinstance(max_val, (int, float)):
            max_score = float(max_val)
        elif isinstance(max_val, str):
            try:
                clean_max = re.sub(r'[^\d.]', '', max_val)
                if clean_max:
                    max_score = float(clean_max)
            except:
                pass
        
        # Calculate percentage if both values are valid
        if score is not None and max_score is not None and max_score > 0:
            percentage = (score / max_score) * 100.0
            return round(percentage, 2)
        
        return None
    
    # Now extract scores for each found dataset
    for dataset_name, row_num in dataset_rows.items():
        # Get score from column AQ (score) and AR (max score)
        score_cell = ws.cell(row=row_num, column=43)  # Column AQ
        max_cell = ws.cell(row=row_num, column=44)    # Column AR
        score_val = score_cell.value
        max_val = max_cell.value
        
        if score_val is None or max_val is None:
            logger.warning(f"No score or max score found for {dataset_name} at row {row_num}")
            continue
        
        # Calculate percentage
        percentage = calculate_percentage(score_val, max_val)
        
        if percentage is None:
            logger.warning(f"Could not calculate percentage for {dataset_name}: score={score_val}, max={max_val}")
            continue
        
        # Determine category
        category = "Other"
        dataset_upper = dataset_name.upper()
        
        if dataset_upper.startswith("HTS"):
            category = "HTS"
        elif dataset_upper.startswith("PREP"):
            category = "PrEP"
        elif "PMTCT" in dataset_upper:
            category = "PMTCT"
        elif dataset_upper.startswith("TB"):
            category = "TB"
        elif dataset_upper.startswith("TX") or "PVLS" in dataset_upper:
            category = "ART"
        elif "POST_RESP" in dataset_upper:
            category = "POST_RESP"  # Not GBV anymore
        
        # Clean display name
        display_name = dataset_name.replace("_", " ").title()
        if "(N)" in display_name:
            display_name = display_name.replace("(N)", "")
        if "(D)" in display_name:
            display_name = display_name.replace("(D)", "")
        display_name = display_name.strip()
        
        thematic_areas.append({
            "name": display_name,
            "original_name": dataset_name,
            "group": category,
            "overall_pct": percentage,
            "row": row_num,
            "raw_score": score_val,
            "max_score": max_val
        })
    
    # ====================================================
    # SPECIAL HANDLING FOR POST_RESP AVERAGE
    # ====================================================
    
    # POST_RESP is average of AQ20, AQ21, AQ22, AQ23
    post_resp_percentages = []
    post_resp_details = []
    
    for row_num in [20, 21, 22, 23]:
        score_val = ws.cell(row=row_num, column=43).value  # Column AQ
        max_val = ws.cell(row=row_num, column=44).value    # Column AR
        
        if score_val is not None and max_val is not None:
            percentage = calculate_percentage(score_val, max_val)
            if percentage is not None:
                post_resp_percentages.append(percentage)
                post_resp_details.append({
                    "row": row_num,
                    "percentage": percentage,
                    "raw_score": score_val,
                    "max_score": max_val
                })
    
    if post_resp_percentages:
        post_resp_avg = sum(post_resp_percentages) / len(post_resp_percentages)
        logger.info(f"POST_RESP average from rows 20-23: {post_resp_avg}% (based on {len(post_resp_percentages)} components)")
        
        # Update or add POST_RESP entry
        post_resp_found = False
        for area in thematic_areas:
            if "POST_RESP" in area["original_name"].upper():
                area["overall_pct"] = round(post_resp_avg, 2)
                area["note"] = f"Average of {len(post_resp_percentages)} components"
                area["components"] = post_resp_details
                post_resp_found = True
                break
        
        if not post_resp_found:
            thematic_areas.append({
                "name": "Post Resp",
                "original_name": "POST_RESP",
                "group": "POST_RESP",
                "overall_pct": round(post_resp_avg, 2),
                "note": f"Average of {len(post_resp_percentages)} components",
                "components": post_resp_details,
                "row": "20-23",
            })
    
    # ====================================================
    # SPECIAL HANDLING FOR TX_ML AVERAGE
    # ====================================================
    
    # TX_ML is average of AQ69, AQ70, AQ71, AQ72, AQ73
    tx_ml_percentages = []
    tx_ml_details = []
    tx_ml_rows = [69, 70, 71, 72, 73]
    
    for row_num in tx_ml_rows:
        score_val = ws.cell(row=row_num, column=43).value  # Column AQ
        max_val = ws.cell(row=row_num, column=44).value    # Column AR
        
        if score_val is not None and max_val is not None:
            percentage = calculate_percentage(score_val, max_val)
            if percentage is not None:
                tx_ml_percentages.append(percentage)
                tx_ml_details.append({
                    "row": row_num,
                    "percentage": percentage,
                    "raw_score": score_val,
                    "max_score": max_val
                })
    
    if tx_ml_percentages:
        tx_ml_avg = sum(tx_ml_percentages) / len(tx_ml_percentages)
        logger.info(f"TX_ML average from rows 69-73: {tx_ml_avg}% (based on {len(tx_ml_percentages)} components)")
        
        # Update or add TX_ML entry
        tx_ml_found = False
        for area in thematic_areas:
            if "TX_ML" in area["original_name"].upper():
                area["overall_pct"] = round(tx_ml_avg, 2)
                area["note"] = f"Average of {len(tx_ml_percentages)} components"
                area["components"] = tx_ml_details
                tx_ml_found = True
                break
        
        if not tx_ml_found:
            thematic_areas.append({
                "name": "TX ML",
                "original_name": "TX_ML",
                "group": "ART",
                "overall_pct": round(tx_ml_avg, 2),
                "note": f"Average of {len(tx_ml_percentages)} components",
                "components": tx_ml_details,
                "row": "69-73",
            })
    
    # ====================================================
    # FILTER OUT UNWANTED DATASETS
    # ====================================================
    
    # Remove datasets we don't want (TX_TB_D, TX_TB_N, TB_ART, TB_PREV_D, TB_PREV_N)
    unwanted_datasets = ["TX_TB_D", "TX_TB_N", "TB_ART", "TB_PREV_D", "TB_PREV_N", "GBV"]
    
    filtered_areas = []
    for area in thematic_areas:
        should_include = True
        for unwanted in unwanted_datasets:
            if unwanted in area["original_name"].upper():
                should_include = False
                logger.info(f"Filtering out: {area['original_name']}")
                break
        
        if should_include:
            filtered_areas.append(area)
    
    thematic_areas = filtered_areas
    
    # Sort thematic areas for consistent display
    thematic_areas.sort(key=lambda x: (
        ["HTS", "POST_RESP", "PMTCT", "TB", "PrEP", "ART"].index(x["group"]) 
        if x["group"] in ["HTS", "POST_RESP", "PMTCT", "TB", "PrEP", "ART"] 
        else 99
    ))
    
    logger.info(f"Extracted {len(thematic_areas)} thematic areas for Q1")
    
    # Log what we found
    for area in thematic_areas[:10]:  # Show first 10
        logger.info(f"  - {area['name']} ({area['group']}): {area['overall_pct']}% (raw: {area.get('raw_score', 'N/A')}/{area.get('max_score', 'N/A')})")
    
    if len(thematic_areas) > 10:
        logger.info(f"  ... and {len(thematic_areas) - 10} more")
    
    # Extract facility info
    facility_info = {
        "facility": facility_name,
        "state": None,
        "lga": None,
        "dqa_date": None,
        "quarter": None
    }
    
    # Create result
    result = {
        "Facility": facility_name,
        "DVV": dvv_pct,
        "Availability": availability,
        "Integrity": integrity,
        "Consistency": consistency,
        "Validity": validity,
        "thematic_areas": thematic_areas,
        "metadata": _extract_file_metadata(dvv_path),
        "facility_for_filtering": facility_name,
        "facility_info": facility_info
    }
    
    # Add aggregated KPIs with updated categories
    if thematic_areas:
        result["thematic_kpis"] = _aggregate_thematic_kpis(thematic_areas)
    
    return result

# -----------------------------
# Thematic KPI Aggregation Function
# -----------------------------

def _aggregate_thematic_kpis(thematic_areas):
    """
    Aggregate thematic areas into KPI categories.
    Includes POST_RESP as separate category (not GBV).
    """
    # Initialize KPI categories
    kpi_data = {
        'ART': {
            'indicators': [],
            'scores': [],
            'average': 0,
            'display_name': 'ART'
        },
        'HTS': {
            'indicators': [],
            'scores': [],
            'average': 0,
            'display_name': 'HTS'
        },
        'PrEP': {
            'indicators': [],
            'scores': [],
            'average': 0,
            'display_name': 'PrEP'
        },
        'PMTCT': {
            'indicators': [],
            'scores': [],
            'average': 0,
            'display_name': 'PMTCT'
        },
        'TB': {
            'indicators': [],
            'scores': [],
            'average': 0,
            'display_name': 'TB'
        },
        'POST_RESP': {  # NEW: Separate category for POST_RESP
            'indicators': [],
            'scores': [],
            'average': 0,
            'display_name': 'POST_RESP'
        }
    }
    
    # Map indicators to KPI categories
    for area in thematic_areas:
        area_name = area['name'].upper()
        score = area['overall_pct']
        
        # ART indicators (TX-related)
        if (area_name.startswith('TX') or 
            'TX_CURR' in area_name or 
            'TX_NEW' in area_name or 
            'TX_PVLS' in area_name or
            'TX_ML' in area_name or
            'TRANSFER' in area_name or
            'TX_RTT' in area_name):
            kpi_data['ART']['indicators'].append(area['name'])
            kpi_data['ART']['scores'].append(score)
        
        # HTS indicators
        elif area_name.startswith('HTS'):
            kpi_data['HTS']['indicators'].append(area['name'])
            kpi_data['HTS']['scores'].append(score)
        
        # POST_RESP indicators
        elif 'POST_RESP' in area_name or 'POST RESP' in area_name:
            kpi_data['POST_RESP']['indicators'].append(area['name'])
            kpi_data['POST_RESP']['scores'].append(score)
        
        # PrEP indicators
        elif area_name.startswith('PREP'):
            kpi_data['PrEP']['indicators'].append(area['name'])
            kpi_data['PrEP']['scores'].append(score)
        
        # PMTCT indicators
        elif 'PMTCT' in area_name:
            kpi_data['PMTCT']['indicators'].append(area['name'])
            kpi_data['PMTCT']['scores'].append(score)
        
        # TB indicators
        elif area_name.startswith('TB'):
            kpi_data['TB']['indicators'].append(area['name'])
            kpi_data['TB']['scores'].append(score)
    
    # Calculate averages
    for category, data in kpi_data.items():
        if data['scores']:
            data['average'] = round(sum(data['scores']) / len(data['scores']), 1)
        else:
            # Remove empty categories
            if category == 'POST_RESP' and not data['scores']:
                del kpi_data['POST_RESP']
                break
    
    return kpi_data


# -----------------------------
# Boolean Sheet Scoring WITH FACILITY FILTERING
# -----------------------------

def _score_boolean_sheet(file_path: str, start_col: str, end_col: str, sheet_type: str = "client", 
                        facility_name: str = None, callback=None): # type: ignore
    """
    TRUE/FALSE scoring with facility filtering.
    """
    _validate_file_extension(file_path)
    
    logger.info(f"=== PROCESSING {sheet_type.upper()} SHEET FOR {facility_name or 'UNKNOWN FACILITY'}: {os.path.basename(file_path)} ===")
    
    metadata = _extract_file_metadata(file_path)
    logger.debug(f"File metadata: {metadata}")
    
    with timer(f"Reading {os.path.basename(file_path)}"):
        df = _read_excel_cached(file_path)
    
    logger.info(f"File has {len(df)} rows, {len(df.columns)} columns")
    
    # Filter by facility if facility_name is provided
    if facility_name:
        df, facility_col_idx = _filter_by_facility(df, facility_name)
        
        if len(df) == 0:
            raise ValueError(f"No data found for facility '{facility_name}' in {file_path}")
    else:
        facility_col_idx = None
        logger.warning("No facility name provided. Processing all rows in file.")
    
    # Find last data row in column C (or facility column if different)
    data_column_idx = facility_col_idx if facility_col_idx is not None else 2
    column_c_data = df.iloc[:, data_column_idx]
    last_data_row = None
    
    for idx in range(len(column_c_data) - 1, -1, -1):
        val = column_c_data.iloc[idx]
        if pd.notna(val):
            if isinstance(val, str):
                if val.strip():
                    last_data_row = idx
                    break
            else:
                last_data_row = idx
                break
    
    if last_data_row is None:
        for idx in range(len(df) - 1, -1, -1):
            row_data = df.iloc[idx]
            if not row_data.isna().all():
                last_data_row = idx
                break
    
    if last_data_row is None:
        raise ValueError("No data found in the sheet after facility filtering")
    
    logger.info(f"Found last data row at index {last_data_row} (Excel row {last_data_row + 2})")
    df = df.iloc[:last_data_row + 1]
    
    # Get appropriate config
    if sheet_type == "vl":
        config = CONFIG["vl_unsuppressed"]
        fallback_start_cols = config.get("fallback_start_cols", [])
        fallback_end_cols = config.get("fallback_end_cols", [])
    else:
        config = CONFIG["client_level"]
        fallback_start_cols = config.get("fallback_start_cols", [])
        fallback_end_cols = config.get("fallback_end_cols", [])
    
    # Find start and end columns
    if sheet_type == "vl":
        logger.info("=== VL SHEET DETECTED ===")
        
        # STRATEGY: Find where the TRUE/FALSE data actually starts
        # We want the simple "Sex" column (not "Sex (Folder)")
        
        # First, try to find the simple "Sex" column (not "Sex (Folder)")
        sex_column_found = None
        for i, col in enumerate(df.columns):
            col_str = str(col).strip().lower()
            if col_str == "sex" or col_str == "sex ":
                # Check if this column has TRUE values
                sample = df[col].dropna().head(10)
                true_count = sum(1 for val in sample if str(val).upper() == "TRUE")
                if true_count >= 5:  # Has at least 5 TRUE values
                    sex_column_found = i
                    logger.info(f"Found TRUE/FALSE 'Sex' column at index {i}: '{df.columns[i]}' with {true_count}/10 TRUE values")
                    break
        
        if sex_column_found is not None:
            start_idx = sex_column_found
            logger.info(f"✓ Using TRUE/FALSE 'Sex' column at index {start_idx}: '{df.columns[start_idx]}'")
        else:
            # Fallback: look for "Sex " with space
            potential_start_indices = []
            for i, col in enumerate(df.columns):
                normalized_col = _normalize_header(col)
                # Look for "sex " with space or similar variations
                if normalized_col.startswith('sex ') or normalized_col == 'sex':
                    potential_start_indices.append((i, col, normalized_col))
            
            if potential_start_indices:
                # Try to find the first one (should be column P)
                start_idx = potential_start_indices[0][0]
                logger.info(f"✓ Using 'Sex ' column at index {start_idx}: '{df.columns[start_idx]}'")
            else:
                # Fallback to search by position
                logger.warning("Could not find 'Sex ' column by name, trying position-based")
                # In the new format, the first "Sex " column should be at position P (16th column, index 15)
                if len(df.columns) > 15:
                    start_idx = 15  # Column P (0-indexed)
                    logger.info(f"✓ Using column position {start_idx} as 'Sex ' (assuming column P)")
                else:
                    start_idx = _find_col_index_with_fallback(df, start_col, fallback_start_cols)
                    logger.info(f"✓ Found start column via fallback at index {start_idx}: '{df.columns[start_idx]}'")
        
        # Find end column (the EAC question)
        # Look for the exact EAC question
        end_patterns = [
            "Does the folder contain the VL result form used to commence EAC?",
            "Does the folder contain the VL result form used to commence EAC",
            "VL result form used to commence EAC",
            "EAC form",
            "eac form"
        ]
        
        end_idx = None
        for pattern in end_patterns:
            try:
                end_idx = _find_col_index(df, pattern)
                logger.info(f"✓ Found end column '{pattern}' at index {end_idx}: '{df.columns[end_idx]}'")
                break
            except ValueError:
                continue
        
        if end_idx is None:
            logger.warning("Could not find EAC column by name, calculating position")
            # From P to AB is 13 columns (P=16, AB=28, 28-16+1=13)
            target_column_count = 13
            end_idx = start_idx + target_column_count - 1
            if end_idx >= len(df.columns):
                end_idx = len(df.columns) - 1
                start_idx = max(0, end_idx - target_column_count + 1)
            
            logger.info(f"Calculated end index: {end_idx} (assuming P to AB is 13 columns)")
    else:
        logger.info("=== CLIENT SHEET DETECTED ===")
        
        # STRATEGY: Find where the TRUE/FALSE data actually starts
        # Based on your data, we want the section that has "Sex" followed by many TRUE values
        
        # First, try to find the simple "Sex" column (not "Sex (Folder)")
        sex_column_found = None
        for i, col in enumerate(df.columns):
            col_str = str(col).strip().lower()
            if col_str == "sex" or col_str == "sex.1":
                # Check if this column has TRUE values
                sample = df[col].dropna().head(10)
                true_count = sum(1 for val in sample if str(val).upper() == "TRUE")
                if true_count >= 5:  # Has at least 5 TRUE values
                    sex_column_found = i
                    logger.info(f"Found TRUE/FALSE 'Sex' column at index {i}: '{df.columns[i]}' with {true_count}/10 TRUE values")
                    break
        
        if sex_column_found is not None:
            start_idx = sex_column_found
            logger.info(f"✓ Using TRUE/FALSE 'Sex' column at index {start_idx}: '{df.columns[start_idx]}'")
        else:
            # Fallback: search by name
            start_idx = _find_col_index_with_fallback(df, start_col, fallback_start_cols)
            logger.info(f"✓ Found start column via name search at index {start_idx}: '{df.columns[start_idx]}'")
        
        # For end column, look for the EXACT match with ".1"
        end_patterns = [
            "Biometric enrollment form available in clients folder.1",
            "biometric enrollment form available in clients folder.1",
            "Biometric enrollment form available",
            "biometric enrollment form"
        ]
        
        end_idx = None
        for pattern in end_patterns:
            try:
                end_idx = _find_col_index(df, pattern)
                logger.info(f"✓ Found end column '{pattern}' at index {end_idx}: '{df.columns[end_idx]}'")
                break
            except ValueError:
                continue
        
        if end_idx is None:
            # Last resort: count 12 columns from start (Q to AB = 12 columns)
            end_idx = start_idx + 11  # Q=17 to AB=28 is 12 columns total
            if end_idx >= len(df.columns):
                end_idx = len(df.columns) - 1
            logger.info(f"Using calculated end index: {end_idx} (12 columns from start)")

    if end_idx < start_idx:
        raise ValueError(f"End column is before start column")

    sub = df.iloc[:, start_idx:end_idx + 1]
    total_rows = len(sub.index)

    if total_rows == 0:
        raise ValueError("No data rows found for TRUE/FALSE scoring after filtering")
    
    logger.info(f"Processing {len(sub.columns)} columns, {total_rows} rows for facility: {facility_name}")
    logger.info(f"Column range: {start_idx} to {end_idx} (Excel columns: {openpyxl.utils.get_column_letter(start_idx + 1)} to {openpyxl.utils.get_column_letter(end_idx + 1)})")

    def to_bool(v):
        """Convert various inputs to boolean"""
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        if pd.isna(v):
            return False
        
        if isinstance(v, (int, float)):
            if v == 1.0 or v == 1:
                return True
            elif v == 0.0 or v == 0:
                return False
            return bool(v)
        
        s = str(v).strip().upper()
        if s == "":
            return False
        
        if s == "1.0" or s == "1":
            return True
        if s == "0.0" or s == "0":
            return False
        
        true_values = {"TRUE", "YES", "Y", "T", "OK", "CHECKED", "PRESENT"}
        return s in true_values

    breakdown = []
    percents = []

    logger.info("=== CALCULATING COLUMN PERCENTAGES ===")
    
    for i, col in enumerate(sub.columns):
        if callback:
            progress = (i + 1) / len(sub.columns) * 100
            callback(progress, f"Processing column {col}")
        
        bools = sub[col].apply(to_bool)
        true_count = int(bools.sum())
        pct = (true_count / total_rows) * 100.0

        breakdown.append({
            "column": str(col),
            "true_count": true_count,
            "total_rows": total_rows,
            "percent": round(pct, 2)
        })
        percents.append(pct)
        
        if i < 3 or i >= len(sub.columns) - 3:
            logger.info(f"Column {i+1}/{len(sub.columns)} '{col}': {true_count}/{total_rows} TRUE ({pct:.2f}%)")

    final_pct = sum(percents) / len(percents) if percents else 0.0
    logger.info(f"Final score for {facility_name} in {file_path}: {final_pct:.2f}%")
    
    if sheet_type == "vl":
        logger.info(f"VL Column percentages ({len(percents)} columns):")
        for i, pct in enumerate(percents):
            logger.info(f"  Column {i}: {pct:.2f}%")
    
    return round(final_pct, 2), breakdown


# -----------------------------
# Public API Functions (UPDATED WITH FACILITY FILTERING)
# -----------------------------

def process_facility_bundle(dvv_filename: str, client_filename: str, vl_filename: str):
    """
    Process the 3 uploads for ONE facility and return ONE result dict.
    Now includes facility filtering for Client Level and VL sheets.
    """
    logger.info(f"Processing facility bundle - DVV: {dvv_filename}, Client: {client_filename}, VL: {vl_filename}")

    dvv_path = os.path.join(UPLOAD_DIR, dvv_filename)
    client_path = os.path.join(UPLOAD_DIR, client_filename)
    vl_path = os.path.join(UPLOAD_DIR, vl_filename)

    for p in [dvv_path, client_path, vl_path]:
        if not os.path.exists(p):
            raise FileNotFoundError(f"Uploaded file not found: {p}")

    # Score #1: DVV extraction (gets facility name)
    with timer("DVV extraction"):
        dvv_result = _extract_dvv_from_template(dvv_path)
    
    dvv_score = dvv_result["DVV"]
    facility_name = dvv_result["Facility"]
    
    logger.info(f"Extracted facility name from DVV: {facility_name}")
    
    # Score #2: Client Level with facility filtering
    with timer("Client Level scoring"):
        client_score, client_breakdown = _score_boolean_sheet(
            client_path,
            start_col="Sex",  # Changed from "Sex.1" to "Sex"
            end_col="Biometric enrollment form available in clients folder.1",
            sheet_type="client",
            facility_name=facility_name  # Pass facility name for filtering
        )

    # Score #3: VL Unsuppressed with facility filtering
    with timer("VL Unsuppressed scoring"):
        vl_score, vl_breakdown = _score_boolean_sheet(
            vl_path,
            start_col="Sex ",
            end_col="Does the folder contain the VL result form used to commence EAC?",
            sheet_type="vl",
            facility_name=facility_name  # Pass facility name for filtering
        )

    # Calculate final score
    final_score = round((dvv_score + client_score + vl_score) / 3.0, 2)

    # Aggregate thematic KPIs for charts
    thematic_kpis = _aggregate_thematic_kpis(dvv_result.get("thematic_areas", []))

    result = {
        **dvv_result,
        "Client_Level": client_score,
        "VL_Unsuppressed": vl_score,
        "Score": final_score,
        "thematic_kpis": thematic_kpis,  # Added aggregated KPIs
        "breakdowns": {
            "client_level": client_breakdown,
            "vl_unsuppressed": vl_breakdown
        },
        "files": {
            "dvv": dvv_filename,
            "client_level": client_filename,
            "vl_unsuppressed": vl_filename
        },
        "processing_timestamp": time.time()
    }

    # Verify scores and assign rating
    warnings = []
    
    if result["DVV"] < 0 or result["DVV"] > 100:
        warnings.append(f"DVV score out of range: {result['DVV']}")
    
    if result["Client_Level"] < 0 or result["Client_Level"] > 100:
        warnings.append(f"Client Level score out of range: {result['Client_Level']}")
    
    if result["VL_Unsuppressed"] < 0 or result["VL_Unsuppressed"] > 100:
        warnings.append(f"VL Unsuppressed score out of range: {result['VL_Unsuppressed']}")
    
    thresholds = CONFIG["thresholds"]
    if result["Score"] >= thresholds["excellent"]:
        result["rating"] = "Excellent"
    elif result["Score"] >= thresholds["good"]:
        result["rating"] = "Good"
    elif result["Score"] >= thresholds["poor"]:
        result["rating"] = "Poor"
    else:
        result["rating"] = "Very Poor"
    
    if warnings:
        result["warnings"] = warnings
        logger.warning(f"Verification warnings for {result['Facility']}: {warnings}")

    logger.info(f"Successfully processed {result['Facility']}. Final score: {final_score}%")
    logger.info(f"Scores breakdown - DVV: {dvv_score}%, Client: {client_score}%, VL: {vl_score}%")
    
    return result


def process_multiple_facilities(facility_list):
    """
    Process multiple facilities in batch.
    """
    logger.info(f"Starting batch processing of {len(facility_list)} facilities")
    
    results = []
    errors = []
    
    for i, facility in enumerate(facility_list):
        logger.info(f"Processing facility {i+1}/{len(facility_list)}: {facility.get('name', 'Unknown')}")
        
        try:
            result = process_facility_bundle(
                facility["dvv"],
                facility["client_level"],
                facility["vl_unsuppressed"]
            )
            results.append(result)
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Failed to process facility {facility.get('name', 'Unknown')}: {error_msg}")
            errors.append({
                "facility": facility.get("name", "Unknown"),
                "error": error_msg,
                "files": facility
            })
    
    logger.info(f"Batch processing complete. Successful: {len(results)}, Failed: {len(errors)}")
    
    return {
        "successful": results,
        "failed": errors,
        "summary": {
            "total": len(facility_list),
            "successful": len(results),
            "failed": len(errors),
            "success_rate": (len(results) / len(facility_list)) * 100 if facility_list else 0
        }
    }


def analyze_dvv_thematic_areas(dvv_filename: str):
    """
    Analyze DVV thematic areas (for the detailed thematic analysis feature).
    """
    logger.info(f"Analyzing DVV thematic areas for: {dvv_filename}")
    
    dvv_path = os.path.join(UPLOAD_DIR, dvv_filename)
    
    if not os.path.exists(dvv_path):
        raise FileNotFoundError(f"DVV file not found: {dvv_path}")
    
    try:
        with timer("DVV thematic area analysis"):
            result = _extract_dvv_from_template(dvv_path)
        
        logger.info(f"Successfully analyzed {len(result.get('thematic_areas', []))} thematic areas")
        return result
        
    except Exception as e:
        logger.error(f"Failed to analyze DVV thematic areas: {e}")
        raise


def compare_multiple_dvv_thematic(facility_files):
    """
    Compare thematic areas across multiple facilities.
    """
    logger.info(f"Comparing thematic areas for {len(facility_files)} facilities")
    
    facilities_data = []
    errors = []
    
    for facility in facility_files:
        try:
            analysis = analyze_dvv_thematic_areas(facility['dvv_file'])
            
            facility_summary = {
                'facility_name': analysis['Facility'],
                'total_indicators': len(analysis.get('thematic_areas', [])),
                'dvv_score': analysis['DVV'],
                'thematic_kpis': analysis.get('thematic_kpis', {}),
                'dvv_file': facility['dvv_file']
            }
            
            facilities_data.append(facility_summary)
            logger.info(f"Processed {facility.get('name', 'Unknown')}: {analysis['DVV']}%")
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Failed to process {facility.get('name', 'Unknown')}: {error_msg}")
            errors.append({
                'facility': facility.get('name', 'Unknown'),
                'error': error_msg
            })
    
    if facilities_data:
        comparative_analysis = {
            'facilities': facilities_data,
            'facilities_count': len(facilities_data),
            'errors': errors
        }
    else:
        comparative_analysis = {
            'facilities': [],
            'errors': errors,
            'message': 'No facilities processed successfully'
        }
    
    return comparative_analysis


def export_results_to_excel(results, output_path):
    """
    Export analysis results to Excel.
    """
    logger.info(f"Exporting {len(results)} results to {output_path}")
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Export summary
            summary_data = []
            for result in results:
                summary_data.append({
                    "Facility": result["Facility"],
                    "DVV Score": result["DVV"],
                    "Client Level": result["Client_Level"],
                    "VL Unsuppressed": result["VL_Unsuppressed"],
                    "Final Score": result["Score"],
                    "Rating": result.get("rating", "N/A"),
                    "Availability": result.get("Availability"),
                    "Integrity": result.get("Integrity"),
                    "Consistency": result.get("Consistency"),
                    "Validity": result.get("Validity")
                })
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)
            
            # Export thematic areas for first few facilities
            max_thematic = min(5, len(results))
            for i in range(max_thematic):
                result = results[i]
                if "thematic_areas" in result and result["thematic_areas"]:
                    thematic_data = []
                    for item in result["thematic_areas"]:
                        thematic_data.append({
                            "Name": item["name"],
                            "Group": item["group"],
                            "Score": item["overall_pct"],
                            "Cell": item["cell"]
                        })
                    
                    if thematic_data:
                        sheet_name = result["Facility"][:25]
                        pd.DataFrame(thematic_data).to_excel(
                            writer,
                            sheet_name=f"{sheet_name}_Thematic",
                            index=False
                        )
            
        logger.info(f"Successfully exported results to {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to export results to Excel: {e}")
        raise


def export_thematic_analysis(thematic_analysis, output_path=None):
    """
    Export thematic analysis results to Excel.
    """
    try:
        if output_path is None:
            facility_name = thematic_analysis.get('Facility', 'Unknown').replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(UPLOAD_DIR, f"{facility_name}_thematic_analysis_{timestamp}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = [
                ["Facility", thematic_analysis.get('Facility', 'Unknown')],
                ["DVV Score", f"{thematic_analysis.get('DVV', 0)}%"],
                ["Availability", f"{thematic_analysis.get('Availability', 0)}%"],
                ["Integrity", f"{thematic_analysis.get('Integrity', 0)}%"],
                ["Consistency", f"{thematic_analysis.get('Consistency', 0)}%"],
                ["Validity", f"{thematic_analysis.get('Validity', 0)}%"],
                ["Total Thematic Areas", len(thematic_analysis.get('thematic_areas', []))]
            ]
            
            pd.DataFrame(summary_data, columns=['Metric', 'Value']).to_excel(
                writer, sheet_name="Summary", index=False
            )
            
            # Thematic areas sheet
            if 'thematic_areas' in thematic_analysis and thematic_analysis['thematic_areas']:
                thematic_data = []
                for item in thematic_analysis['thematic_areas']:
                    thematic_data.append({
                        "Indicator": item["name"],
                        "Category": item["group"],
                        "Score": f"{item['overall_pct']}%",
                        "Cell Reference": item["cell"]
                    })
                
                pd.DataFrame(thematic_data).to_excel(
                    writer, sheet_name="Thematic Areas", index=False
                )
            
            # KPI summary sheet
            if 'thematic_kpis' in thematic_analysis:
                kpi_data = []
                for category, data in thematic_analysis['thematic_kpis'].items():
                    kpi_data.append({
                        "KPI Category": category,
                        "Average Score": f"{data.get('average', 0)}%",
                        "Indicators Count": len(data.get('indicators', []))
                    })
                
                pd.DataFrame(kpi_data).to_excel(
                    writer, sheet_name="KPI Summary", index=False
                )
        
        logger.info(f"Thematic analysis exported to: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Failed to export thematic analysis: {e}")
        raise